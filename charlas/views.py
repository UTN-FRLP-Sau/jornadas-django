import threading
from django.utils import timezone
from django import forms as django_forms
from .models import Registration, Talk, Certificate, CertificateConfig
from weasyprint import HTML
from django.template.loader import render_to_string
from django.core.mail import get_connection
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import uuid
import csv
import io
import json
from io import BytesIO
import re

import qrcode
from django.conf import settings
from django.core.mail import EmailMessage
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from functools import wraps
from django.contrib.auth import views as auth_views

from charlas.constants import DEPT_COLORS, DEPT_NAMES, DEPT_ORDER
from .forms import RegistrationForm, TalkForm
from .models import Registration, Talk, CertificateConfig, EmissionJob, Certificate, TalkRating, Survey

import openpyxl
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
import google.generativeai as genai

# ────────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────────


def _analizar_respuestas_ia(respuestas, contexto):
    try:
        genai.configure(api_key=settings.GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-1.5-flash')

        prompt = f"""Analizá las siguientes respuestas abiertas de una encuesta sobre {contexto} 
de las Jornadas de Formación Profesional de la UTN La Plata 2026.

Respuestas:
{chr(10).join(f'- {r}' for r in respuestas if r.strip())}

Respondé en español con este formato exacto:
**Temas principales:** (listado de los temas más mencionados)
**Aspectos positivos:** (lo que más valoran)
**Aspectos a mejorar:** (críticas y sugerencias)
**Resumen:** (2-3 oraciones resumiendo el sentimiento general)"""

        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f'Error al analizar: {e}'


def _generate_certificate_pdf(cert):
    template_path = settings.BASE_DIR / 'charlas' / 'static' / \
        'charlas' / 'img' / 'template_certificado.pdf'
    output_path = settings.MEDIA_ROOT / \
        'certificados' / f'cert_{cert.codigo}.pdf'
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Crear overlay con el texto
    overlay_buffer = BytesIO()
    c = canvas.Canvas(overlay_buffer, pagesize=landscape(A4))
    width, height = landscape(A4)  # 842 x 595 pts

    # Apellido — mayúsculas, centrado
    c.setFont('Helvetica-Bold', 28)
    c.setFillColor(colors.HexColor('#0F172B'))
    apellido = cert.apellido.upper()
    c.drawCentredString(width / 2, height / 2 + 20, apellido)

    # Nombre — title case
    c.setFont('Helvetica', 24)
    nombre = cert.nombre.title()
    c.drawCentredString(width / 2, height / 2 - 20, nombre)

    # DNI — más pequeño
    c.setFont('Helvetica', 14)
    c.setFillColor(colors.HexColor('#666565'))
    c.drawCentredString(width / 2, height / 2 - 55, f'DNI: {cert.dni}')

    # Frase de validación — abajo, pequeña
    c.setFont('Helvetica', 9)
    c.setFillColor(colors.HexColor('#999999'))
    validate_url = f'{settings.SITE_URL}/certificado/validar/'
    c.drawCentredString(
        width / 2, 60, f'Validá este certificado ingresando el código {cert.codigo} en: {validate_url}')

    c.save()
    overlay_buffer.seek(0)

    # Superponer sobre el template
    template_pdf = PdfReader(str(template_path))
    overlay_pdf = PdfReader(overlay_buffer)

    writer = PdfWriter()
    page = template_pdf.pages[0]
    page.merge_page(overlay_pdf.pages[0])
    writer.add_page(page)

    with open(output_path, 'wb') as f:
        writer.write(f)

    return f'certificados/cert_{cert.codigo}.pdf'



def _run_emission(job_id, config_id):
    import django
    from charlas.models import Registration, Certificate, CertificateConfig, EmissionJob

    job = EmissionJob.objects.get(id=job_id)
    config = CertificateConfig.objects.get(id=config_id)
    job.status = 'procesando'
    job.save()

    try:
        dnis = Registration.objects.filter(
            attended=True).values_list('dni', flat=True).distinct()
        elegibles = [dni for dni in dnis if _evaluar_alumno(
            dni, config) and not Certificate.objects.filter(dni=dni).exists()]
        job.total = len(elegibles)
        job.save()

        for dni in elegibles:
            reg = Registration.objects.filter(
                dni=dni, attended=True).select_related('talk').first()
            cert = Certificate.objects.create(
                nombre=reg.nombre,
                apellido=reg.apellido,
                dni=reg.dni,
                legajo=reg.legajo,
                correo=reg.correo,
                config=config,
            )
            ok = _send_certificate_email(cert)
            if ok:
                job.enviados += 1
            else:
                job.errores += 1
            job.save()

        job.status = 'completado'
        job.finished_at = timezone.now()
        job.save()

    except Exception as e:
        job.status = 'error'
        job.finished_at = timezone.now()
        job.save()
        print(f'[EMISSION] Error: {e}')


def _send_certificate_email(cert):
    try:
        from django.core.mail import get_connection
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from django.template.loader import render_to_string

        html_body = render_to_string('charlas/email_certificado.html', {
            'cert': cert,
            'site_url': settings.SITE_URL,
        })

        msg = MIMEMultipart('mixed')
        msg['Subject'] = 'Tu certificado — Jornadas de Formación Profesional 2026'
        msg['From'] = settings.DEFAULT_FROM_EMAIL
        msg['To'] = cert.correo

        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText(html_body, 'html'))
        msg.attach(alt)
        '''
        # Generar PDF
        archivo_path = settings.MEDIA_ROOT / cert.archivo if cert.archivo else None

        if archivo_path and archivo_path.exists():
            with open(archivo_path, 'rb') as f:
                pdf_bytes = f.read()

            from email.mime.application import MIMEApplication
            pdf_mime = MIMEApplication(pdf_bytes, _subtype='pdf')
            pdf_mime.add_header(
                'Content-Disposition', 'attachment',
                filename=f'certificado_{cert.apellido}_{cert.nombre}.pdf'
            )
            msg.attach(pdf_mime)
        '''

        connection = get_connection()
        connection.open()
        connection.connection.sendmail(
            settings.DEFAULT_FROM_EMAIL,
            [cert.correo],
            msg.as_string()
        )
        connection.close()
        return True
    except Exception as exc:
        print(f'[CERT] Error enviando a {cert.correo}: {exc}')
        return False


def _evaluar_alumno(dni, config):
    """
    Devuelve True si el alumno cumple las condiciones de la config.
    """
    regs = Registration.objects.filter(
        dni=dni, attended=True).select_related('talk')

    if not regs.exists():
        return False

    if config.requiere_magistral:
        tiene_magistral = regs.filter(talk__department='Magistral').exists()
        if not tiene_magistral:
            return False

    # Excluir magistrales del conteo
    regs_no_magistral = regs.exclude(talk__department='Magistral')

    if config.modalidad == 'total':
        return regs_no_magistral.count() >= config.minimo
    
    elif config.modalidad == 'por_dia':
        dias = {}
        for reg in regs_no_magistral:
            fecha = reg.talk.date
            dias.setdefault(fecha, 0)
            dias[fecha] += 1
        return all(count >= config.minimo for count in dias.values()) and len(dias) > 0

    return False

def sanitize_excel(value):
    if not isinstance(value, str):
        return value
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)


def login_required_json(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'message': 'No autenticado.'}, status=401)
        return view_func(request, *args, **kwargs)
    return wrapper


def _build_qr_content(registration):
    return f"{registration.apellido}|{registration.legajo}|{registration.dni}|{registration.talk_id}|{registration.token}"


def _send_confirmation_email(request, registration):
    talk = registration.talk
    cancel_url = f"{settings.SITE_URL}/cancel/{registration.token}/"

    # Generar QR
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(_build_qr_content(registration))
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    qr_bytes = buf.getvalue()

    html_body = f"""
    <h3>¡Hola {registration.nombre}!</h3>
    <p>Tu inscripción a la charla <b>{talk.title}</b> ha sido confirmada.</p>
    <p><b>Fecha:</b> {talk.date} | <b>Hora:</b> {talk.time} | <b>Disertante:</b> {talk.speaker}</p>
    <p>Para asistir el día del evento, presentá el siguiente código QR en la entrada:</p>
    <img src="cid:qr_code" alt="QR Asistencia" />
    <br><br>
    <p>Si no puedes asistir, cancelá tu inscripción para liberar el cupo:</p>
    <p><a href="{cancel_url}">{cancel_url}</a></p>
    """

    try:
        msg = MIMEMultipart('related')
        msg['Subject'] = f'Confirmación de Inscripción: {talk.title}'
        msg['From'] = settings.DEFAULT_FROM_EMAIL
        msg['To'] = registration.correo

        alt = MIMEMultipart('alternative')
        msg.attach(alt)
        alt.attach(MIMEText(html_body, 'html'))

        img_mime = MIMEImage(qr_bytes)
        img_mime.add_header('Content-ID', '<qr_code>')
        img_mime.add_header('Content-Disposition', 'inline',
                            filename='qr_asistencia.png')
        msg.attach(img_mime)

        connection = get_connection()
        connection.open()
        connection.connection.sendmail(
            settings.DEFAULT_FROM_EMAIL,
            [registration.correo],
            msg.as_string()
        )
        connection.close()
        return True
    except Exception as exc:
        print(f'[EMAIL] No se pudo enviar a {registration.correo}: {exc}')
        return False


# ────────────────────────────────────────────────────────────────────────────────
# Public views
# ────────────────────────────────────────────────────────────────────────────────

def index(request):
    talks = list(Talk.objects.all())
    config = CertificateConfig.objects.filter(activa=True).first()
    descarga_habilitada = config.descarga_habilitada if config else False
    
    def get_date_priority(talk):
        d = talk.date.lower()
        if 'martes' in d: return 1
        if 'miércoles' in d or 'miercoles' in d: return 2
        if 'jueves' in d: return 3
        return 4
        
    def normalize_date(talk):
        d = talk.date.lower()
        if 'martes' in d: return 'Martes 19 de Mayo'
        if 'miércoles' in d or 'miercoles' in d: return 'Miércoles 20 de Mayo'
        if 'jueves' in d: return 'Jueves 21 de Mayo'
        return talk.date

    for t in talks:
        t.normalized_date = normalize_date(t)
        
    # Ordenar por prioridad de fecha (Martes, Miercoles, Jueves) y luego por hora
    sorted_talks = sorted(talks, key=lambda t: (get_date_priority(t), t.time))
    
    return render(request, 'charlas/index.html', {
        'talks': sorted_talks,
        'descarga_habilitada': descarga_habilitada,
        })

def _duplicate_exists(talk, dni, legajo):
    """Check if a registration with the same DNI or legajo already exists for this talk."""
    from django.db.models import Q
    return Registration.objects.filter(talk=talk).filter(
        Q(dni=dni) | Q(legajo=legajo)
    ).exists()


# Rewrite talk_detail to use _duplicate_exists properly
def talk_register(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    form = RegistrationForm()
    errors = []

    if request.method == 'POST':
        if talk.remaining_capacity <= 0:
            errors = ['Los cupos para esta charla están agotados.']
        else:
            form = RegistrationForm(request.POST)
            if form.is_valid():
                data = form.cleaned_data
                if _duplicate_exists(talk, data['dni'], data['legajo']):
                    errors = [
                        'Ya te encontrás inscripto en esta charla con ese DNI o Legajo.']
                else:
                    token = str(uuid.uuid4())
                    reg = Registration.objects.create(
                        talk=talk,
                        nombre=data['nombre'],
                        apellido=data['apellido'],
                        dni=data['dni'],
                        legajo=data['legajo'],
                        correo=data['correo'],
                        token=token,
                    )
                    email_ok = _send_confirmation_email(request, reg)
                    return render(request, 'charlas/success.html', {
                        'talk': talk,
                        'email_ok': email_ok,
                    })
            else:
                for field_errors in form.errors.values():
                    errors.extend(field_errors)

    return render(request, 'charlas/talk.html', {'talk': talk, 'form': form, 'errors': errors})


def cancel_registration(request, token):
    reg = get_object_or_404(Registration, token=token)
    reg.delete()
    return render(request, 'charlas/cancel_success.html')


# ────────────────────────────────────────────────────────────────────────────────
# Admin views
# ────────────────────────────────────────────────────────────────────────────────

@login_required
def admin_dashboard(request):
    talks = Talk.objects.all()
    return render(request, 'charlas/admin_dashboard.html', {'talks': talks})


""" 
@login_required
def scanner_dashboard(request):
    talks = Talk.objects.all()
    return render(request, 'charlas/scanner_dashboard.html', {'talks': talks})
 """

@login_required
def admin_new_talk(request):
    form = TalkForm()
    if request.method == 'POST':
        form = TalkForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    return render(request, 'charlas/talk_form.html', {
        'form': form,
        'title': 'Crear Nueva Charla',
        'submit_label': 'Crear Charla',
    })


@login_required
def admin_edit_talk(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    form = TalkForm(instance=talk)
    if request.method == 'POST':
        form = TalkForm(request.POST, request.FILES, instance=talk)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    return render(request, 'charlas/talk_form.html', {
        'form': form,
        'talk': talk,
        'title': 'Editar Charla',
        'submit_label': 'Guardar Cambios',
        'current_image': talk.image if talk.image else None,
    })


@login_required
@require_POST
def admin_delete_talk(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    talk.delete()
    return redirect('admin_dashboard')


@login_required
def admin_register_student(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    form = RegistrationForm()
    errors = []

    if request.method == 'POST':
        if talk.remaining_capacity <= 0:
            errors = ['Los cupos para esta charla están agotados.']
        else:
            form = RegistrationForm(request.POST)
            if form.is_valid():
                data = form.cleaned_data
                if _duplicate_exists(talk, data['dni'], data['legajo']):
                    errors = ['Ya se encuentra inscripto en esta charla con ese DNI o Legajo.']
                else:
                    token = str(uuid.uuid4())
                    reg = Registration.objects.create(
                        talk=talk,
                        nombre=data['nombre'],
                        apellido=data['apellido'],
                        dni=data['dni'],
                        legajo=data['legajo'],
                        correo=data['correo'],
                        token=token,
                        attended=True,
                    )
                    _send_confirmation_email(request, reg)
                    return redirect('admin_talk_details', pk=talk.pk)
            else:
                for field_errors in form.errors.values():
                    errors.extend(field_errors)

    return render(request, 'charlas/talk.html', {'talk': talk, 'form': form, 'errors': errors})


@login_required
def admin_talk_details(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    registrations = talk.registrations.all().order_by('apellido', 'nombre')
    return render(request, 'charlas/admin_talk_details.html', {
        'talk': talk,
        'registrations': registrations,
    })


@login_required
@require_POST
def update_attendance(request, reg_id):
    import json
    reg = get_object_or_404(Registration, pk=reg_id)
    data = json.loads(request.body)
    reg.attended = data.get('attended', False)
    reg.save()
    return JsonResponse({'success': True})


@login_required
@require_POST
def admin_delete_registration(request, reg_id):
    reg = get_object_or_404(Registration, pk=reg_id)
    talk_id = reg.talk_id
    reg.delete()
    return redirect('admin_talk_details', pk=talk_id)


@login_required
def admin_scan(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    return render(request, 'charlas/admin_scan.html', {'talk': talk})


@login_required_json
@require_POST
def api_scan(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

    import json
    data = json.loads(request.body)
    token = data.get('token')
    talk_id = data.get('talk_id')

    if not token or not talk_id:
        return JsonResponse({'success': False, 'message': 'Datos incompletos.'}, status=400)

    try:
        token = data.get('token').split('|')[-1]
        reg = Registration.objects.get(token=token, talk_id=talk_id)
    except Registration.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Código QR inválido para esta charla.'}, status=404)

    if reg.attended:
        return JsonResponse({
            'success': True,
            'message': f'{reg.nombre} {reg.apellido} ya tenía asistencia registrada.',
            'already_attended': True,
        })

    reg.attended = True
    reg.save()
    return JsonResponse({
        'success': True,
        'message': f'Asistencia confirmada para {reg.nombre} {reg.apellido}.',
        'already_attended': False,
    })


@login_required
def export_attendance(request, pk):
    talk = get_object_or_404(Talk, pk=pk)
    registrations = talk.registrations.filter(attended=True).order_by('apellido', 'nombre')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asistencias'
    headers = ['Apellido', 'Nombre', 'DNI', 'Legajo', 'Correo']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for reg in registrations:
        ws.append([reg.apellido, reg.nombre, reg.dni, reg.legajo, reg.correo])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Asistencias_{talk.title.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_talks(request):
    talks = Talk.objects.all().order_by('department', 'date', 'time')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja default

    departments = talks.values_list(
        'department', flat=True).distinct().order_by('department')

    for dept in departments:
        ws = wb.create_sheet(title=dept[:31])  # Excel limita a 31 chars
        headers = ['Título', 'Fecha', 'Hora', 'Lugar', 'Disertante',
                   'Descripcion']
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for talk in talks.filter(department=dept):
            ws.append([
                talk.title,
                talk.date,
                talk.time,
                talk.location or 'A confirmar',
                talk.speaker,
                talk.description or '',
            ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="charlas_por_departamento.xlsx"'
    return response

PDF_CRONOGRAMA_PATH = settings.BASE_DIR / 'media' / 'cronograma_jfp2026.pdf'

@login_required
def export_cronograma_pdf(request):
    regenerar = request.GET.get('regenerar') == '1'
    
    if not PDF_CRONOGRAMA_PATH.exists() or regenerar:
        talks = Talk.objects.all().order_by('date', 'time')
        departments = {dept: {} for dept in DEPT_ORDER}
        
        for talk in talks:
            dept = talk.department
            if dept not in departments:
                departments[dept] = {}
            if talk.date not in departments[dept]:
                departments[dept][talk.date] = []
            departments[dept][talk.date].append(talk)

        # Ordenar fechas dentro de cada depto
        DATE_ORDER = ['Martes 19 de Mayo', 'Miércoles 20 de Mayo', 'Jueves 21 de Mayo']
        for dept in departments:
            departments[dept] = {
                fecha: departments[dept][fecha]
                for fecha in DATE_ORDER
                if fecha in departments[dept]
            }
        
        depts_list = [
            (dept, DEPT_NAMES.get(dept, dept), fechas, DEPT_COLORS.get(dept, '#2b4efe'))
            for dept, fechas in departments.items()
            if fechas
        ]

        html_string = render_to_string('charlas/cronograma_pdf.html', {
            'departments': departments,
            'depts_list': depts_list,
            'dept_names': DEPT_NAMES,
        })
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
        PDF_CRONOGRAMA_PATH.write_bytes(pdf)

    return FileResponse(
        open(PDF_CRONOGRAMA_PATH, 'rb'),
        content_type='application/pdf',
        as_attachment=True,
        filename='cronograma_jfp2026.pdf'
    )


@login_required
def import_attendance(request):
    from .forms import AttendanceImportForm
    form = AttendanceImportForm()

    if request.method == 'POST':
        form = AttendanceImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            talk_destino = form.cleaned_data['talk']
            decoded = csv_file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded), delimiter=';')

            results = []
            for row in reader:
                raw = row['NOMBRE'].strip('"')
                parts = raw.split('|')
                if len(parts) < 5:
                    results.append({
                        'token': raw,
                        'talk_id': None,
                        'nombre': '-',
                        'charla': '-',
                        'estado': 'Fila inválida',
                        'raw': raw
                    })
                    continue

                apellido, legajo, dni, talk_id, token = parts[0], parts[1], parts[2], parts[3], parts[4]

                try:
                    reg = Registration.objects.select_related('talk').get(
                        talk=talk_destino, dni=dni
                    )
                    if reg.attended:
                        estado = 'Ya presente'
                    else:
                        reg.attended = True
                        reg.save()
                        estado = 'Actualizado'
                    results.append({
                        'token': token,
                        'talk_id': talk_destino.id,
                        'nombre': f"{reg.apellido}, {reg.nombre}",
                        'charla': reg.talk.title,
                        'estado': estado,
                        'raw': raw,
                    })
                except Registration.DoesNotExist:
                    results.append({
                        'token': token,
                        'talk_id': talk_destino.id,
                        'nombre': apellido,
                        'charla': '-',
                        'estado': 'No encontrado',
                        'raw': raw,
                    })

            talk_ids = set(r['talk_id'] for r in results if r.get('talk_id'))
            talk_ref = Talk.objects.filter(
                pk__in=talk_ids).first() if talk_ids else None

            request.session['import_results'] = results

            return render(request, 'charlas/import_attendance.html', {
                'form': form,
                'results': results,
                'talk_ref': talk_destino,
                'summary': {
                    'actualizados':   sum(1 for r in results if r['estado'] == 'Actualizado'),
                    'ya_presentes':   sum(1 for r in results if r['estado'] == 'Ya presente'),
                    'no_encontrados': sum(1 for r in results if r['estado'] == 'No encontrado'),
                }
            })

    return render(request, 'charlas/import_attendance.html', {'form': form})


@login_required
def export_import_results(request):
    results = request.session.get('import_results', [])
    if not results:
        return redirect('import_attendance')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resultados'
    headers = ['Nombre', 'Charla', 'Estado', 'Línea CSV']
    ws.append(headers)
    for col in range(1, 4):
        ws.cell(row=1, column=col).font = Font(bold=True)

    colors = {
        'Actualizado':   'C6EFCE',
        'Ya presente':   'FFEB9C',
        'No encontrado': 'FFC7CE',
        'Fila inválida': 'D9D9D9',
    }

    for r in results:
        ws.append([
            sanitize_excel(r['nombre']),
            sanitize_excel(r['charla']),
            sanitize_excel(r['estado']),
            sanitize_excel(r.get('raw', '—'))
        ])
        fill_color = colors.get(r['estado'], 'FFFFFF')
        for col in range(1, 4):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type='solid'
            )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="resultado_importacion.xlsx"'
    return response


@login_required
def certificate_dashboard(request):
    config = CertificateConfig.objects.filter(activa=True).first()
    elegibles = []

    if config:
        # Obtener todos los DNIs con asistencia
        dnis = Registration.objects.filter(
            attended=True).values_list('dni', flat=True).distinct()

        for dni in dnis:
            if _evaluar_alumno(dni, config):
                reg = Registration.objects.filter(
                    dni=dni, attended=True).select_related('talk').first()
                ya_tiene = Certificate.objects.filter(dni=dni).exists()
                elegibles.append({
                    'dni': dni,
                    'nombre': reg.nombre,
                    'apellido': reg.apellido,
                    'legajo': reg.legajo,
                    'correo': reg.correo,
                    'charlas': Registration.objects.filter(dni=dni, attended=True).count(),
                    'ya_emitido': ya_tiene,
                })

    return render(request, 'charlas/certificate_dashboard.html', {
        'config': config,
        'elegibles': elegibles,
        'total': len(elegibles),
        'pendientes': sum(1 for e in elegibles if not e['ya_emitido']),
    })


@login_required
def certificate_config(request):
    config = CertificateConfig.objects.filter(activa=True).first()

    class ConfigForm(django_forms.ModelForm):
        class Meta:
            model = CertificateConfig
            fields = ['modalidad', 'minimo', 'requiere_magistral', 'descarga_habilitada', 'mensaje_bloqueado']
            widgets = {
                'modalidad': django_forms.Select(attrs={'class': 'form-select'}),
                'minimo': django_forms.NumberInput(attrs={'class': 'form-control', 'min': 1}),
                'requiere_magistral': django_forms.CheckboxInput(attrs={'class': 'form-check-input'}),
                'descarga_habilitada': django_forms.CheckboxInput(attrs={'class': 'form-check-input'}),
                'mensaje_bloqueado': django_forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
            }
            labels = {
                'modalidad': 'Modalidad',
                'minimo': 'Mínimo de charlas',
                'requiere_magistral': 'Requiere al menos una magistral',
                'descarga_habilitada': 'Habilitar descarga de certificados',
                'mensaje_bloqueado': 'Mensaje cuando la descarga está bloqueada',
            }

    form = ConfigForm(instance=config)

    if request.method == 'POST':
        form = ConfigForm(request.POST, instance=config)
        if form.is_valid():
            # Desactivar configs anteriores
            CertificateConfig.objects.update(activa=False)
            nueva = form.save(commit=False)
            nueva.activa = True
            nueva.save()
            return redirect('certificate_dashboard')

    return render(request, 'charlas/certificate_config.html', {'form': form, 'config': config})


@login_required
def certificate_emit(request):
    config = CertificateConfig.objects.filter(activa=True).first()
    if not config:
        return redirect('certificate_dashboard')

    dnis = Registration.objects.filter(
        attended=True).values_list('dni', flat=True).distinct()
    emitidos = 0
    errores = 0

    for dni in dnis:
        if not _evaluar_alumno(dni, config):
            continue
        if Certificate.objects.filter(dni=dni).exists():
            continue

        reg = Registration.objects.filter(
            dni=dni, attended=True).select_related('talk').first()

        cert = Certificate.objects.create(
            nombre=reg.nombre,
            apellido=reg.apellido,
            dni=reg.dni,
            legajo=reg.legajo,
            correo=reg.correo,
            config=config,
        )
        try:
            archivo = _generate_certificate_pdf(cert)
            cert.archivo = archivo
            cert.save()
        except Exception as e:
            print(f'[CERT PDF] Error generando PDF para {cert.dni}: {e}')

        # Generar y enviar — por ahora placeholder
        ok = _send_certificate_email(cert)
        if ok:
            emitidos += 1
        else:
            errores += 1

    return render(request, 'charlas/certificate_emit_result.html', {
        'emitidos': emitidos,
        'errores': errores,
    })


@login_required
def certificate_emit(request):
    config = CertificateConfig.objects.filter(activa=True).first()
    if not config:
        return redirect('certificate_dashboard')

    # Si hay un job procesando, redirigir al status
    job_activo = EmissionJob.objects.filter(status='procesando').first()
    if job_activo:
        return redirect('certificate_emit_status', job_id=job_activo.id)

    if request.method == 'POST':
        job = EmissionJob.objects.create()
        t = threading.Thread(target=_run_emission, args=(
            job.id, config.id), daemon=True)
        t.start()
        return redirect('certificate_emit_status', job_id=job.id)

    return render(request, 'charlas/certificate_emit_confirm.html', {'config': config})


@login_required
def certificate_emit_status(request, job_id):
    job = get_object_or_404(EmissionJob, id=job_id)
    return render(request, 'charlas/certificate_emit_status.html', {'job': job})


@login_required
def certificate_emit_status_api(request, job_id):
    job = get_object_or_404(EmissionJob, id=job_id)
    return JsonResponse({
        'status': job.status,
        'total': job.total,
        'enviados': job.enviados,
        'errores': job.errores,
        'finished': job.status in ('completado', 'error'),
    })


def certificate_validate(request):
    cert = None
    regs = []
    error = None

    if request.method == 'POST':
        dni = request.POST.get('dni', '').strip()
        codigo = request.POST.get('codigo', '').strip()

        try:
            cert = Certificate.objects.get(dni=dni, codigo=codigo)
            regs = Registration.objects.filter(
                dni=dni, attended=True
            ).select_related('talk').order_by('talk__date', 'talk__time')
        except Certificate.DoesNotExist:
            error = 'No se encontró un certificado con ese DNI y código.'

    return render(request, 'charlas/certificate_validate.html', {
        'cert': cert,
        'regs': regs,
        'error': error,
    })


def certificate_download(request):
    config = CertificateConfig.objects.filter(activa=True).first()
     # Verificar si la descarga está habilitada
    if not config or not config.descarga_habilitada:
        mensaje = config.mensaje_bloqueado if config else 'La descarga de certificados no está disponible.'
        return render(request, 'charlas/certificate_blocked.html', {'mensaje': mensaje})

    cert = None
    regs = []
    cumple = False
    error = None

    if request.method == 'POST':
        dni = request.POST.get('dni', '').strip()

        if not dni:
            error = 'Ingresá tu DNI.'
        else:
            cert = Certificate.objects.filter(dni=dni).first()
            if cert:
                cumple = True
                # Generar PDF si no existe
                if not cert.archivo:
                    try:
                        archivo = _generate_certificate_pdf(cert)
                        cert.archivo = archivo
                        cert.save()
                    except Exception as e:
                        print(f'[CERT PDF] Error: {e}')
                # Si no completó la encuesta, redirigir
                survey_obj = Survey.objects.filter(certificate=cert).first()
                if not survey_obj or not survey_obj.completada:
                    return redirect('survey', dni=dni)
                regs = Registration.objects.filter(
                    dni=dni, attended=True
                ).select_related('talk').order_by('talk__date', 'talk__time')
            else:
                regs = Registration.objects.filter(
                    dni=dni, attended=True
                ).select_related('talk').order_by('talk__date', 'talk__time')
                cumple = False

    return render(request, 'charlas/certificate_download.html', {
        'cert': cert,
        'regs': regs,
        'cumple': cumple,
        'error': error,
        'submitted': request.method == 'POST',
    })


def survey(request, dni, step=1):
    cert = get_object_or_404(Certificate, dni=dni)
    survey_obj, _ = Survey.objects.get_or_create(certificate=cert)

    if survey_obj.completada:
        return redirect('survey_done', dni=dni)

    regs = Registration.objects.filter(
        dni=dni, attended=True
    ).select_related('talk').order_by('talk__date', 'talk__time')

    # paso 1: bienvenida, 2: datos generales, 3: evaluación general
    # 4: feria empresas, 5: feria laboratorios, 6+N: charlas, último: final
    total_steps = 5 + regs.count() + 1  # +1 para la página final

    if request.method == 'POST':
        if step == 2:
            survey_obj.carrera = request.POST.get('carrera', '')
            survey_obj.anio_cursada = request.POST.get('anio_cursada', '')
            survey_obj.save()

        elif step == 3:
            survey_obj.evaluacion_general = request.POST.get(
                'evaluacion_general', '')
            survey_obj.aporte_formacion = request.POST.get(
                'aporte_formacion', '')
            survey_obj.interes_tematicas = request.POST.get(
                'interes_tematicas', '')
            survey_obj.organizacion_general = request.POST.get(
                'organizacion_general') or None
            survey_obj.matriz_variedad = request.POST.get(
                'matriz_variedad', '')
            survey_obj.matriz_disertantes = request.POST.get(
                'matriz_disertantes', '')
            survey_obj.matriz_horarios = request.POST.get(
                'matriz_horarios', '')
            survey_obj.matriz_informacion = request.POST.get(
                'matriz_informacion', '')
            survey_obj.matriz_inscripcion = request.POST.get(
                'matriz_inscripcion', '')
            survey_obj.matriz_acreditacion = request.POST.get(
                'matriz_acreditacion', '')
            survey_obj.matriz_espacios = request.POST.get(
                'matriz_espacios', '')
            survey_obj.matriz_colaboradores = request.POST.get(
                'matriz_colaboradores', '')
            survey_obj.lo_mejor = request.POST.get('lo_mejor', '')
            survey_obj.a_mejorar = request.POST.get('a_mejorar', '')
            survey_obj.save()

        elif step == 4:
            survey_obj.asistio_feria_empresas = request.POST.get(
                'asistio_feria_empresas') == 'si'
            if survey_obj.asistio_feria_empresas:
                survey_obj.evaluacion_feria_empresas = request.POST.get(
                    'evaluacion_feria_empresas', '')
                survey_obj.utilidad_feria_empresas = request.POST.get(
                    'utilidad_feria_empresas', '')
                survey_obj.stands_interesantes = request.POST.get(
                    'stands_interesantes', '')
                survey_obj.mejoras_feria_empresas = request.POST.get(
                    'mejoras_feria_empresas', '')
            survey_obj.save()

        elif step == 5:
            survey_obj.asistio_feria_laboratorios = request.POST.get(
                'asistio_feria_laboratorios') == 'si'
            if survey_obj.asistio_feria_laboratorios:
                survey_obj.evaluacion_feria_laboratorios = request.POST.get(
                    'evaluacion_feria_laboratorios', '')
                survey_obj.conocio_proyectos = request.POST.get(
                    'conocio_proyectos', '')
                survey_obj.lab_interesante = request.POST.get(
                    'lab_interesante', '')
                survey_obj.mejoras_feria_laboratorios = request.POST.get(
                    'mejoras_feria_laboratorios', '')
            survey_obj.save()

        elif step >= 6 and step <= 5 + regs.count():
            talk_index = step - 6
            if talk_index < regs.count():
                reg = regs[talk_index]
                TalkRating.objects.update_or_create(
                    survey=survey_obj,
                    talk=reg.talk,
                    defaults={
                        'puntuacion_disertante': request.POST.get('puntuacion_disertante'),
                        'puntuacion_contenido': request.POST.get('puntuacion_contenido'),
                        'comentario': request.POST.get('comentario', ''),
                    }
                )

        elif step == total_steps:
            survey_obj.proxima_edicion = request.POST.get(
                'proxima_edicion', '')
            survey_obj.completada = True
            survey_obj.save()
            return redirect('survey_done', dni=dni)

        return redirect('survey_step', dni=dni, step=step + 1)

    context = {
        'cert': cert,
        'step': step,
        'total_steps': total_steps,
        'survey': survey_obj,
    }

    if step == 1:
        template = 'charlas/survey_welcome.html'
    elif step == 2:
        template = 'charlas/survey_datos.html'
    elif step == 3:
        context['matriz_items'] = [
            ('matriz_variedad', 'Variedad de temáticas'),
            ('matriz_disertantes', 'Nivel de exposición de los disertantes'),
            ('matriz_horarios', 'Organización de horarios'),
            ('matriz_informacion', 'Información brindada previamente'),
            ('matriz_inscripcion', 'Sistema de inscripción'),
            ('matriz_acreditacion', 'Sistema de acreditación de asistencia'),
            ('matriz_espacios', 'Distribución de aulas y espacios'),
            ('matriz_colaboradores', 'Acompañamiento de colaboradores'),
        ]
        template = 'charlas/survey_general.html'
    elif step == 4:
        template = 'charlas/survey_empresas.html'
    elif step == 5:
        template = 'charlas/survey_laboratorios.html'
    elif step >= 6 and step <= 5 + regs.count():
        talk_index = step - 6
        context['reg'] = regs[talk_index]
        context['talk_num'] = talk_index + 1
        template = 'charlas/survey_talk.html'
    elif step == total_steps:
        template = 'charlas/survey_final.html'
    else:
        return redirect('survey', dni=dni)

    return render(request, template, context)


def survey_done(request, dni):
    cert = get_object_or_404(Certificate, dni=dni)
    if not cert.archivo:
        try:
            from charlas.views import _generate_certificate_pdf
            archivo = _generate_certificate_pdf(cert)
            cert.archivo = archivo
            cert.save()
        except Exception as e:
            print(f'[CERT PDF] Error: {e}')
    return render(request, 'charlas/survey_done.html', {'cert': cert})


@login_required
def attendance_dashboard(request):
    from charlas.constants import DEPT_COLORS, DEPT_ORDER

    total_inscriptos = Registration.objects.values('dni').distinct().count()
    total_presentes = Registration.objects.filter(
        attended=True).values('dni').distinct().count()

    departments = {}
    for dept in DEPT_ORDER:
        talks = Talk.objects.filter(department=dept)
        inscriptos_unicos = Registration.objects.filter(
            talk__department=dept
        ).values('dni').distinct().count()
        presentes_unicos = Registration.objects.filter(
            talk__department=dept, attended=True
        ).values('dni').distinct().count()
        dept_data = {
            'dept': dept,
            'color': DEPT_COLORS.get(dept, '#2b4efe'),
            'inscriptos': inscriptos_unicos,
            'presentes': presentes_unicos,
            'charlas': []
        }
        for talk in talks:
            dept_data['charlas'].append({
                'titulo': talk.title,
                'fecha': talk.date,
                'inscriptos': talk.registered_count,
                'presentes': talk.registrations.filter(attended=True).count(),
            })
        departments[dept] = dept_data

    data = list(departments.values())
    total_charlas = sum(len(d['charlas']) for d in data)

    return render(request, 'charlas/attendance_dashboard.html', {
        'departments': json.dumps(data),
        'total_inscriptos': total_inscriptos,
        'total_presentes': total_presentes,
        'total_charlas': total_charlas,
    })


@login_required
def survey_dashboard(request):
    carreras = Survey.objects.exclude(carrera='').values_list(
        'carrera', flat=True).distinct()
    anios = Survey.objects.exclude(anio_cursada='').values_list(
        'anio_cursada', flat=True).distinct()
    total = Survey.objects.filter(completada=True).count()

    return render(request, 'charlas/survey_dashboard.html', {
        'carreras': sorted(carreras),
        'anios': sorted(anios),
        'total': total,
    })


@login_required
def survey_dashboard_api(request):
    carrera = request.GET.get('carrera', '')
    anio = request.GET.get('anio', '')

    qs = Survey.objects.filter(completada=True)
    if carrera:
        qs = qs.filter(carrera=carrera)
    if anio:
        qs = qs.filter(anio_cursada=anio)

    total = qs.count()
    if total == 0:
        return JsonResponse({'total': 0})

    # Evaluación general
    def dist(field):
        from django.db.models import Count
        return dict(qs.exclude(**{f'{field}': ''}).values_list(field).annotate(n=Count('id')))

    # Promedio organización
    from django.db.models import Avg
    org_avg = qs.filter(organizacion_general__isnull=False).aggregate(
        avg=Avg('organizacion_general'))['avg']

    # Matriz
    matriz_fields = [
        ('matriz_variedad', 'Variedad de temáticas'),
        ('matriz_disertantes', 'Nivel de disertantes'),
        ('matriz_horarios', 'Organización de horarios'),
        ('matriz_informacion', 'Información previa'),
        ('matriz_inscripcion', 'Sistema de inscripción'),
        ('matriz_acreditacion', 'Sistema de acreditación'),
        ('matriz_espacios', 'Distribución de espacios'),
        ('matriz_colaboradores', 'Acompañamiento de colaboradores'),
    ]
    orden_matriz = ['Muy bueno', 'Bueno', 'Regular', 'Malo', 'Muy malo']
    matriz_data = []
    for field, label in matriz_fields:
        d = dist(field)
        matriz_data.append({
            'label': label,
            'valores': {v: d.get(v, 0) for v in orden_matriz}
        })

    # Ferias
    asistio_empresas = qs.filter(asistio_feria_empresas=True).count()
    asistio_laboratorios = qs.filter(asistio_feria_laboratorios=True).count()

    # Charlas — promedio por charla
    from django.db.models import Avg as AvgF
    talk_ratings = TalkRating.objects.filter(survey__in=qs).values(
        'talk__title'
    ).annotate(
        avg=AvgF('puntuacion_disertante')
    ).order_by('-avg')

    return JsonResponse({
        'total': total,
        'evaluacion_general': dist('evaluacion_general'),
        'aporte_formacion': dist('aporte_formacion'),
        'interes_tematicas': dist('interes_tematicas'),
        'organizacion_avg': round(org_avg, 2) if org_avg else None,
        'matriz': matriz_data,
        'asistio_empresas': asistio_empresas,
        'asistio_laboratorios': asistio_laboratorios,
        'evaluacion_feria_empresas': dist('evaluacion_feria_empresas'),
        'evaluacion_feria_laboratorios': dist('evaluacion_feria_laboratorios'),
        'talk_ratings': list(talk_ratings),
    })


@login_required
def survey_analizar_ia(request):
    carrera = request.GET.get('carrera', '')
    anio = request.GET.get('anio', '')
    campo = request.GET.get('campo', 'lo_mejor')

    qs = Survey.objects.filter(completada=True)
    if carrera:
        qs = qs.filter(carrera=carrera)
    if anio:
        qs = qs.filter(anio_cursada=anio)

    campos_validos = ['lo_mejor', 'a_mejorar', 'stands_interesantes',
                      'mejoras_feria_empresas', 'lab_interesante',
                      'mejoras_feria_laboratorios', 'proxima_edicion']

    if campo not in campos_validos:
        return JsonResponse({'error': 'Campo inválido'}, status=400)

    respuestas = list(qs.exclude(**{campo: ''}).values_list(campo, flat=True))

    if not respuestas:
        return JsonResponse({'resultado': 'No hay respuestas para analizar.'})

    contexto_map = {
        'lo_mejor': 'lo que más gustó del evento',
        'a_mejorar': 'aspectos a mejorar del evento',
        'stands_interesantes': 'stands de la feria de empresas',
        'mejoras_feria_empresas': 'mejoras para la feria de empresas',
        'lab_interesante': 'laboratorios más interesantes',
        'mejoras_feria_laboratorios': 'mejoras para la feria de laboratorios',
        'proxima_edicion': 'sugerencias para la próxima edición',
    }

    resultado = _analizar_respuestas_ia(
        respuestas, contexto_map.get(campo, campo))
    return JsonResponse({'resultado': resultado, 'total_respuestas': len(respuestas)})


@login_required
def survey_export(request):
    carrera = request.GET.get('carrera', '')
    anio = request.GET.get('anio', '')

    qs = Survey.objects.filter(completada=True).select_related('certificate')
    if carrera:
        qs = qs.filter(carrera=carrera)
    if anio:
        qs = qs.filter(anio_cursada=anio)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Encuestas'
    headers = [
        'DNI', 'Nombre', 'Carrera', 'Año', 'Evaluación General',
        'Aporte Formación', 'Interés Temáticas', 'Organización (1-5)',
        'Lo mejor', 'A mejorar', 'Asistió Feria Empresas',
        'Evaluación Feria Empresas', 'Asistió Feria Laboratorios',
        'Evaluación Feria Laboratorios', 'Próxima Edición'
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for s in qs:
        ws.append([
            s.certificate.dni,
            f"{s.certificate.apellido}, {s.certificate.nombre}",
            s.carrera, s.anio_cursada,
            s.evaluacion_general, s.aporte_formacion, s.interes_tematicas,
            s.organizacion_general,
            s.lo_mejor, s.a_mejorar,
            'Sí' if s.asistio_feria_empresas else 'No',
            s.evaluacion_feria_empresas,
            'Sí' if s.asistio_feria_laboratorios else 'No',
            s.evaluacion_feria_laboratorios,
            s.proxima_edicion,
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="encuestas.xlsx"'
    return response
